import os
import openpyxl
import random
from crud.invite_code import create_invite_code
from crud.club import create_club
from database import SessionLocal, Base, engine

os.environ.setdefault('DATABASE_URL', 'sqlite:////www/bupt-bingo/Server/bingo.db')

def init_club(file_path):
    """
    输入的 Excel 文件应包含社团名单（社工摊位也算作一个社团），第一行为表头，第一列为序号，第二列为社团名称，第三列为社团类型（是否为五佳十优社团）（0 或 1）
    该函数将为每个社团创建社团记录
    """
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    # 加载xlsx文件（支持写入）
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # 获取第一个工作表
    
    # 遍历行（从第二行开始，min_row=2）
    # 注意：iter_rows返回的row是单元格元组，row[0]是第一列，row[1]是第二列...
    for row in sheet.iter_rows(min_row=2):
        if len(row) < 3:  # 检查当前行是否至少有3列
            continue
        print(row[0].value, row[1].value, row[2].value)  # 打印第一列、第二列和第三列内容
        col2_value = row[1].value
        col3_value = row[2].value
        if not col2_value:
            continue
        if not col3_value:
            continue
        # 创建社团
        club = create_club(db, club_name=col2_value, club_type=col3_value)
        print(f"创建社团: id={club.id}, name={club.club_name}")
    # 保存并关闭
    workbook.save(file_path)
    workbook.close()
    db.close()

if __name__ == "__main__":
    file_path = "参与社团名单.xlsx"
    init_club(file_path)